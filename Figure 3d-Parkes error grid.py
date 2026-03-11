from pathlib import Path
from collections import Counter

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.drawing.image import Image as XLImage
import matplotlib.pyplot as plt

MMOL_TO_MGDL = 18.0182
ZONE_LABELS = {0: "A", 1: "B", 2: "C", 3: "D", 4: "E"}


def slope(x1, y1, x2, y2):
    if x2 == x1:
        raise ValueError("vertical line")
    return (y2 - y1) / (x2 - x1)


def y_at_x(x1, y1, x, m):
    return (x - x1) * m + y1


def x_at_y(x1, y1, y, m):
    return (y - y1) / m + x1


def point_in_polygon(x, y, polygon):
    inside = False
    n = len(polygon)

    for i in range(n):
        x1, y1 = polygon[i]
        x2, y2 = polygon[(i + 1) % n]

        cross = (x - x1) * (y2 - y1) - (y - y1) * (x2 - x1)
        if abs(cross) < 1e-9:
            if min(x1, x2) - 1e-9 <= x <= max(x1, x2) + 1e-9 and min(y1, y2) - 1e-9 <= y <= max(y1, y2) + 1e-9:
                return True

        intersects = ((y1 > y) != (y2 > y))
        if intersects:
            xinters = (x2 - x1) * (y - y1) / (y2 - y1 + 1e-12) + x1
            if x < xinters:
                inside = not inside

    return inside


def build_parkes_polygons(diabetes_type, n=1.0, max_x=18, max_y=18):
    if diabetes_type == 1:
        ce = slope(35, 155, 50, 550)
        cdu = slope(80, 215, 125, 550)
        cdl = slope(250, 40, 550, 150)
        ccu = slope(70, 110, 260, 550)
        ccl = slope(260, 130, 550, 250)
        cbu = slope(280, 380, 430, 550)
        cbl = slope(385, 300, 550, 450)

        limit_e = [
            (0, 150 / n),
            (35 / n, 155 / n),
            (x_at_y(35 / n, 155 / n, max_y, ce), max_y),
            (0, max_y),
        ]

        limit_d_lower = [
            (250 / n, 0),
            (250 / n, 40 / n),
            (max_x, y_at_x(410 / n, 110 / n, max_x, cdl)),
            (max_x, 0),
        ]

        limit_d_upper = [
            (0, 100 / n),
            (25 / n, 100 / n),
            (50 / n, 125 / n),
            (80 / n, 215 / n),
            (x_at_y(80 / n, 215 / n, max_y, cdu), max_y),
            (0, max_y),
        ]

        limit_c_lower = [
            (120 / n, 0),
            (120 / n, 30 / n),
            (260 / n, 130 / n),
            (max_x, y_at_x(260 / n, 130 / n, max_x, ccl)),
            (max_x, 0),
        ]

        limit_c_upper = [
            (0, 60 / n),
            (30 / n, 60 / n),
            (50 / n, 80 / n),
            (70 / n, 110 / n),
            (x_at_y(70 / n, 110 / n, max_y, ccu), max_y),
            (0, max_y),
        ]

        limit_b_lower = [
            (50 / n, 0),
            (50 / n, 30 / n),
            (170 / n, 145 / n),
            (385 / n, 300 / n),
            (max_x, y_at_x(385 / n, 300 / n, max_x, cbl)),
            (max_x, 0),
        ]

        limit_b_upper = [
            (0, 50 / n),
            (30 / n, 50 / n),
            (140 / n, 170 / n),
            (280 / n, 380 / n),
            (x_at_y(280 / n, 380 / n, max_y, cbu), max_y),
            (0, max_y),
        ]

        return [
            (limit_e, 4),
            (limit_d_lower, 3),
            (limit_d_upper, 3),
            (limit_c_lower, 2),
            (limit_c_upper, 2),
            (limit_b_lower, 1),
            (limit_b_upper, 1),
        ]

    elif diabetes_type == 2:
        ce = slope(35, 200, 50, 550)
        cdu = slope(35, 90, 125, 550)
        cdl = slope(410, 110, 550, 160)
        ccu = slope(30, 60, 280, 550)
        ccl = slope(260, 130, 550, 250)
        cbu = slope(230, 330, 440, 550)
        cbl = slope(330, 230, 550, 450)

        limit_e = [
            (0, 200 / n),
            (35 / n, 200 / n),
            (x_at_y(35 / n, 200 / n, max_y, ce), max_y),
            (0, max_y),
        ]

        limit_d_lower = [
            (250 / n, 0),
            (250 / n, 40 / n),
            (410 / n, 110 / n),
            (max_x, y_at_x(410 / n, 110 / n, max_x, cdl)),
            (max_x, 0),
        ]

        limit_d_upper = [
            (0, 80 / n),
            (25 / n, 80 / n),
            (35 / n, 90 / n),
            (x_at_y(35 / n, 90 / n, max_y, cdu), max_y),
            (0, max_y),
        ]

        limit_c_lower = [
            (90 / n, 0),
            (260 / n, 130 / n),
            (max_x, y_at_x(260 / n, 130 / n, max_x, ccl)),
            (max_x, 0),
        ]

        limit_c_upper = [
            (0, 60 / n),
            (30 / n, 60 / n),
            (x_at_y(30 / n, 60 / n, max_y, ccu), max_y),
            (0, max_y),
        ]

        limit_b_lower = [
            (50 / n, 0),
            (50 / n, 30 / n),
            (90 / n, 80 / n),
            (330 / n, 230 / n),
            (max_x, y_at_x(330 / n, 230 / n, max_x, cbl)),
            (max_x, 0),
        ]

        limit_b_upper = [
            (0, 50 / n),
            (30 / n, 50 / n),
            (230 / n, 330 / n),
            (x_at_y(230 / n, 330 / n, max_y, cbu), max_y),
            (0, max_y),
        ]

        return [
            (limit_e, 4),
            (limit_d_lower, 3),
            (limit_d_upper, 3),
            (limit_c_lower, 2),
            (limit_c_upper, 2),
            (limit_b_lower, 1),
            (limit_b_upper, 1),
        ]

    else:
        raise ValueError("diabetes_type must be 1 or 2")


def classify_parkes(ref_mgdl, pred_mgdl, diabetes_type):
    max_x = max(550.0, ref_mgdl + 20.0, pred_mgdl + 20.0)
    max_y = max(550.0, ref_mgdl + 20.0, pred_mgdl + 20.0)
    polygons = build_parkes_polygons(diabetes_type, n=1.0, max_x=max_x, max_y=max_y)

    zone = 0
    for poly, z in polygons:
        if point_in_polygon(ref_mgdl, pred_mgdl, poly):
            zone = z
            break

    return zone, ZONE_LABELS[zone]


def make_summary(df, zone_col):
    counts = Counter(df[zone_col].dropna().tolist())
    total = sum(counts.values())
    rows = []
    for z in ["A", "B", "C", "D", "E"]:
        c = counts.get(z, 0)
        pct = (c / total * 100) if total else 0
        rows.append((z, c, pct))
    return total, rows


def draw_parkes_plot(df, diabetes_type, out_png):
    fig, ax = plt.subplots(figsize=(7.5, 7.5))

    x = df["BG_mM"].to_numpy()
    y = df["pBG"].to_numpy()

    ax.scatter(x, y, s=14, alpha=0.8)
    ax.plot([0, 18], [0, 18], linestyle=":")

    if diabetes_type == 1:
        lines = [
            ([0, 30], [50, 50]),
            ([30, 140], [50, 170]),
            ([140, 280], [170, 380]),
            ([280, 430], [380, 550]),
            ([0, 50], [60, 80]),
            ([50, 70], [80, 110]),
            ([70, 260], [110, 550]),
            ([120, 120], [0, 30]),
            ([120, 260], [30, 130]),
            ([260, 550], [130, 250]),
            ([0, 25], [100, 100]),
            ([25, 50], [100, 125]),
            ([50, 80], [125, 215]),
            ([80, 125], [215, 550]),
            ([250, 250], [0, 40]),
            ([250, 550], [40, 150]),
            ([0, 35], [150, 155]),
            ([35, 50], [155, 550]),
        ]
    else:
        lines = [
            ([0, 30], [50, 50]),
            ([30, 230], [50, 330]),
            ([230, 440], [330, 550]),
            ([0, 30], [60, 60]),
            ([30, 280], [60, 550]),
            ([90, 260], [0, 130]),
            ([260, 550], [130, 250]),
            ([0, 25], [80, 80]),
            ([25, 35], [80, 90]),
            ([35, 125], [90, 550]),
            ([250, 250], [0, 40]),
            ([250, 410], [40, 110]),
            ([410, 550], [110, 160]),
            ([0, 35], [200, 200]),
            ([35, 50], [200, 550]),
        ]

    for xs, ys in lines:
        xs_mm = [v / MMOL_TO_MGDL for v in xs]
        ys_mm = [v / MMOL_TO_MGDL for v in ys]
        ax.plot(xs_mm, ys_mm, linewidth=1)

    ax.set_xlim(0, 18)
    ax.set_ylim(0, 18)
    ax.set_aspect("equal", adjustable="box")
    ax.set_xlabel("Measured BG (mM)")
    ax.set_ylabel("Personalized BG (mM)")
    ax.set_title(f"Parkes Error Grid - Type {diabetes_type}")

    plt.tight_layout()
    fig.savefig(out_png, dpi=180, bbox_inches="tight")
    plt.close(fig)


def style_sheet(ws):
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for c in col:
            value = "" if c.value is None else str(c.value)
            max_len = max(max_len, len(value))
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 12), 28)


def write_dataframe(ws, df):
    for row in dataframe_to_rows(df, index=False, header=True):
        ws.append(row)
    style_sheet(ws)


def create_summary_sheet(ws, df, diabetes_type, source_filename, img_path):
    ws["A1"] = f"Parkes Error Grid Summary - Type {diabetes_type}"
    ws["A1"].font = Font(bold=True, size=14)

    zone_col = f"Type{diabetes_type}_Zone"
    total, rows = make_summary(df, zone_col)
    a_count = dict((z, c) for z, c, _ in rows).get("A", 0)

    ws["A3"] = "Source file"
    ws["B3"] = source_filename
    ws["A4"] = "Rows analyzed"
    ws["B4"] = total
    ws["A5"] = "Units"
    ws["B5"] = "Input in mmol/L; Parkes grid calculated in mg/dL (x18.0182)"
    ws["A6"] = "Result"
    ws["B6"] = f"{a_count} / {total} in Zone A"

    ws["A8"] = "Zone"
    ws["B8"] = "Count"
    ws["C8"] = "Percent"

    for i, (z, c, p) in enumerate(rows, start=9):
        ws[f"A{i}"] = z
        ws[f"B{i}"] = c
        ws[f"C{i}"] = p / 100
        ws[f"C{i}"].number_format = "0.00%"

    for cell in ["A8", "B8", "C8"]:
        ws[cell].fill = PatternFill("solid", fgColor="1F4E78")
        ws[cell].font = Font(color="FFFFFF", bold=True)

    img = XLImage(str(img_path))
    img.width = 520
    img.height = 520
    ws.add_image(img, "E2")

    for col in ["A", "B", "C", "E", "F", "G", "H", "I", "J"]:
        ws.column_dimensions[col].width = 16
    ws.row_dimensions[1].height = 24


def generate_parkes_workbook(input_xlsx, output_xlsx=None):
    input_xlsx = Path(input_xlsx)

    if output_xlsx is None:
        output_xlsx = input_xlsx.with_name(f"{input_xlsx.stem}_Parkes_error_grid.xlsx")
    else:
        output_xlsx = Path(output_xlsx)

    original_df = pd.read_excel(input_xlsx, sheet_name=0)

    required_cols = {"BG_mM", "pBG", "filename"}
    missing = required_cols - set(original_df.columns)
    if missing:
        raise ValueError(f"Missing required columns: {sorted(missing)}")

    df = original_df.copy()
    df["BG_mg_dL"] = df["BG_mM"] * MMOL_TO_MGDL
    df["pBG_mg_dL"] = df["pBG"] * MMOL_TO_MGDL

    t1_detailed = []
    t1_zone = []
    t2_detailed = []
    t2_zone = []

    for ref, pred in zip(df["BG_mg_dL"], df["pBG_mg_dL"]):
        z1_num, z1_txt = classify_parkes(ref, pred, diabetes_type=1)
        z2_num, z2_txt = classify_parkes(ref, pred, diabetes_type=2)

        t1_detailed.append(z1_num)
        t1_zone.append(z1_txt)
        t2_detailed.append(z2_num)
        t2_zone.append(z2_txt)

    df["Type1_Detailed_Zone"] = t1_detailed
    df["Type1_Zone"] = t1_zone
    df["Type2_Detailed_Zone"] = t2_detailed
    df["Type2_Zone"] = t2_zone

    type1_png = input_xlsx.with_name("parkes_type1.png")
    type2_png = input_xlsx.with_name("parkes_type2.png")
    draw_parkes_plot(df, 1, type1_png)
    draw_parkes_plot(df, 2, type2_png)

    wb = Workbook()
    default_ws = wb.active
    wb.remove(default_ws)

    ws_raw = wb.create_sheet("Sheet1")
    write_dataframe(ws_raw, original_df)

    ws_data = wb.create_sheet("Parkes_Data")
    write_dataframe(ws_data, df)

    ws_sum1 = wb.create_sheet("Summary_Type1")
    create_summary_sheet(
        ws_sum1,
        df,
        diabetes_type=1,
        source_filename=input_xlsx.name,
        img_path=type1_png,
    )

    ws_sum2 = wb.create_sheet("Summary_Type2")
    create_summary_sheet(
        ws_sum2,
        df,
        diabetes_type=2,
        source_filename=input_xlsx.name,
        img_path=type2_png,
    )

    ws_notes = wb.create_sheet("Notes")
    notes = [
        ["Method notes", ""],
        ["Input format", "Original sheet preserved; analysis added in new sheets."],
        ["Conversion", "Parkes error grids are defined in mg/dL. mmol/L values were converted using x18.0182."],
        ["Type 1/Type 2", "Both Parkes Type 1 and Type 2 summaries are included because diabetes type was not specified."],
    ]
    for row in notes:
        ws_notes.append(row)

    ws_notes["A1"].font = Font(bold=True, size=14)
    ws_notes.column_dimensions["A"].width = 18
    ws_notes.column_dimensions["B"].width = 95

    wb.save(output_xlsx)
    print(f"Saved: {output_xlsx}")


if __name__ == "__main__":
    generate_parkes_workbook(
        input_xlsx=r"",
        output_xlsx=r""
    )